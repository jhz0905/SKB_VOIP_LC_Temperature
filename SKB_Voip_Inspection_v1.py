import os
import re
from openpyxl import load_workbook

print("### Wait.. Inspection Ongoing ###")
nowdir = os.getcwd()
logdir = "\log"
nextdir = nowdir+logdir
os.chdir(str(nextdir))

file_list = os.listdir(os.getcwd())



### Parsing ###

for file_num in range(len(file_list)):
    os_name = 0
    hostname = 0
    hw_type = 0
    tem_line1 = 0
    tem_line2 = 0
    pid_line1 = 0
    pid_line2 = 0
    tem_num = 0
    high_lcnum = 0
    high_lcpid = 0
    compare_result = 0
    lc_threshold = 0
    used_slot = ""
    lc_slot = list()
    lc_tem = list()
    lc_name = list()

    log_file = open("%s" %file_list[file_num], "r")
    log_line = log_file.readlines()
    log_file.close()

    ### OS ###
    for os_parsing in range(len(log_line)):
        os_pattern = r'^Cisco\s(IOS|Internetwork Operating System|IOS XR)\sSoftware'
        os_result = re.search(os_pattern, log_line[os_parsing])
        if os_result != None:
            os_name = os_result.group()
            break
        else:
            pass

    ### Hostname ###
    if os_name == "Cisco IOS XR Software":
        for host_parsing in range(len(log_line)):
            host_pattern = r'^\w\w/\d/\w\w\d/\w\w\w\d:.*#show version brief$'
            host_result = re.search(host_pattern, log_line[host_parsing])
            if host_result != None:
                hostname = host_result.group()
                hostname = hostname.split(sep = ":", maxsplit = 1)
                hostname = hostname[1]
                hostname = hostname.split(sep = "#", maxsplit = 1)
                hostname = hostname[0]
                hw_type = "ASR-9912"
                break
            else:
                pass
        os.rename("%s" %file_list[file_num],"%s.txt" %hostname)
        #print("Filename %s -> %s.txt" %(file_list[file_num], hostname))
    else:
        for host_parsing in range(len(log_line)):
            host_pattern = r'^.*#show version$'
            host_result = re.search(host_pattern, log_line[host_parsing])
            if host_result != None:
                hostname = host_result.group()
                hostname = hostname.split(sep = "#s", maxsplit = 1)
                hostname = hostname[0]
                break
            else:
                pass
        os.rename("%s" %file_list[file_num],"%s.txt" %hostname)
        #print("Filename %s -> %s.txt" %(file_list[file_num], hostname))

        for hw_parsing in range(len(log_line)):
            hw_pattern = r'^PID:\s.*\s*.*$'
            hw_result = re.search(hw_pattern, log_line[hw_parsing])
            if hw_result != None:
                hw_type = hw_result.group()
                hw_type = hw_type.split(sep = " ", maxsplit = 2)
                hw_type = hw_type[1]
                break
            else:
                pass

### Part. ASR9912 ###    

    if hw_type == "ASR-9912":
        for tem_start1 in range(len(log_line)):
            tem_pattern1 = r'^\w\w/\d/\w\w\d/\w\w\w\d:.*#show environment temperatures$'
            tem_result1 = re.search(tem_pattern1, log_line[tem_start1])
            if tem_result1 != None:
                tem_line1 = tem_start1
                break
            else:
                pass

        for tem_start2 in range(len(log_line)):
            tem_pattern2 = r'^\w\w/\d/\w\w\d/\w\w\w\d:.*#show environment table$'
            tem_result2 = re.search(tem_pattern2, log_line[tem_start2])
            if tem_result2 != None:
                tem_line2 = tem_start2
                break
            else:
                pass

        for search_lc in range(tem_line1, tem_line2):
            lc_pattern = r'^\d/\d/.$'
            lc_result = re.search(lc_pattern, log_line[search_lc])
            if lc_result != None:
                lc_list = lc_result.group()
                lc_list = lc_list.split(sep = "/", maxsplit = 2)
                lc_slot.append(lc_list[1])
            else:
                pass

        for search_tem in range(tem_line1, tem_line2):
            lctem_pattern = r'^\s*host\s*Hotspot0\s*\d\d.\d$'
            lctem_result = re.search(lctem_pattern, log_line[search_tem])
            if lctem_result != None: 
                lctem_list = lctem_result.group()
                lctem_pattern = r'\d{2,3}.\d'
                lctem_result2 = re.search(lctem_pattern, lctem_list)
                lctem_list = lctem_result2.group()
                #lctem_list = lctem_list.split(sep = " ")
                #lctem_list = lctem_list[28]
                if tem_num <= 1:
                    pass
                else:
                    lc_tem.append(lctem_list)
                tem_num = tem_num + 1
            else:
                pass
        
        for pid_start in range(len(log_line)):
            pid_pattern = r'^\w*/\d/\w*\d/\w*\d:.*#admin show (plat|platform)$'
            pid_result = re.search(pid_pattern, log_line[pid_start])
            if pid_result != None:
                pid_line1 = pid_start
                break
            else:
                pass
        
        for pid_stop in range(len(log_line)):
            pid_pattern = r'^\w*/\d/\w*\d/\w*\d:.*#show environment temperatures$'
            pid_result = re.search(pid_pattern, log_line[pid_stop])
            if pid_result != None:
                pid_line2 = pid_stop
                break
            else:
                pass

        for search_pid in range(pid_line1, pid_line2):
            pid_pattern = r'^\d/\d/\w*\d\s*\S*'
            pid_result = re.search(pid_pattern, log_line[search_pid])
            if pid_result != None:
                pid_result2 = pid_result.group()
                pid_pattern = r'\w+-\w+-\w*'
                pid_name = re.search(pid_pattern, pid_result2) 
                pid_name = pid_name.group() 
                lc_name.append(pid_name)          
            else:
                pass            

        for tem_compare in range(0,len(lc_tem)):
            if str(compare_result) < lc_tem[tem_compare]:
                compare_result = lc_tem[tem_compare]
                high_lcnum = lc_slot[tem_compare]
                high_lcpid = lc_name[tem_compare]

        #print(hostname)
        #print("%s %s %s" %(high_lcnum, high_lcpid, compare_result))

### Part. C4510 ###

    elif hw_type == "WS-C4510R":
        for search_tem in range(len(log_line)):
            tem_pattern = r'^Chassis Temperature\s*=\s\d*\sdegrees\sCelsius'
            tem_result = re.search(tem_pattern, log_line[search_tem])
            if tem_result != None:
                tem_pattern = r'\d\d'
                tem_result2 = tem_result.group()
                tem_regex = re.search(tem_pattern, tem_result2)
                compare_result = tem_regex.group()
                high_lcnum = "Chassis"
                high_lcpid = "Chassis"
                break
            else:
                pass
        #print(hostname)
        #print("%s %s %s" %(high_lcnum, high_lcpid, compare_result))

### Part. C6509 & C7609 ###

    else:
        for search_pid in range(len(log_line)):
            pid_pattern = r'^\s+[^56]\s+\d+\s+.*\w+-\w+-\w+\s+\w+$'
            pid_result = re.search(pid_pattern, log_line[search_pid])
            if pid_result != None:
                pid_result2 = pid_result.group()
                pid_pattern = r'\w+-\w+-\w+'
                pid_name = re.search(pid_pattern, pid_result2) 
                pid_name = pid_name.group()
                lc_name.append(pid_name)
            else:
                pass            

        for search_tem in range(len(log_line)):
            tem_pattern = r'^module [^56] outlet temperature: \d\dC$'
            tem_result = re.search(tem_pattern, log_line[search_tem])
            if tem_result != None:
                tem_result2 = tem_result.group()
                num_pattern = r'\d'
                tem_pattern = r'\d{2,3}'
                num_regex = re.search(num_pattern, tem_result2)
                tem_regex = re.search(tem_pattern, tem_result2)
                num_regex = num_regex.group()
                tem_regex = tem_regex.group()
                lc_slot.append(num_regex)
                lc_tem.append(tem_regex)
            else:
                pass
        
        for tem_compare in range(0,len(lc_tem)):
            if str(compare_result) < lc_tem[tem_compare]:
                compare_result = lc_tem[tem_compare]
                high_lcnum = lc_slot[tem_compare]
                high_lcpid = lc_name[tem_compare]
        
        for search_thres in range(len(log_line)):
            thres_pattern = r'^\s+threshold #2 for module %s outlet temperature:\s$' %high_lcnum
            thres_result = re.search(thres_pattern, log_line[search_thres])
            if thres_result != None:
                thres_pattern = r'\d+'
                thres_result2 = re.search(thres_pattern, log_line[search_thres + 1])
                lc_threshold = thres_result2.group()
                break
            else:
                pass

        #print(hostname)
        #print("%s %s %s %s" %(high_lcnum, high_lcpid, compare_result, lc_threshold))

    for append_lc in range(len(lc_slot)):
        used_slot = "%s %s," %(used_slot, lc_slot[append_lc])
    used_lc_pattern = r'(\d, )+\d'
    used_lc_result = re.search(used_lc_pattern, used_slot)
    if used_lc_result != None:
        used_lc_result = used_lc_result.group()
    else:
        pass

    #print(hostname)
    #print(used_lc_result)

    wb = load_workbook("%s\SKB VoIP 온도 조사.xlsx" %nowdir)
    ws = wb["카드별 온도점검"]
    for excel_row in range(2,25):
        if ws["D%s" %excel_row].value == hostname:
            if hw_type == "WS-C6509-E" or hw_type == "CISCO7609-S":
                ws["E%s" %excel_row] = high_lcnum
                ws["F%s" %excel_row] = compare_result
                ws["H%s" %excel_row] = lc_threshold
                ws["I%s" %excel_row] = used_lc_result
                ws["J%s" %excel_row] = high_lcpid
            else:
                ws["E%s" %excel_row] = high_lcnum
                ws["F%s" %excel_row] = compare_result
                ws["I%s" %excel_row] = used_lc_result
                ws["J%s" %excel_row] = high_lcpid
        else:
            pass
    
    wb.save("%s\SKB VoIP 온도 조사.xlsx" %nowdir)

print("### Inspection Successful ###")
input("Press enter to exit ;)")
