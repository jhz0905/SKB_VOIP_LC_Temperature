import os
import re
from openpyxl import load_workbook

print("### Wait.. Inspection Ongoing ###")
nowdir = os.getcwd()
logdir = "\log"
nextdir = nowdir+logdir
os.chdir(str(nextdir))

file_list = os.listdir(os.getcwd())



### Parsing ###

for file_num in range(len(file_list)):
    os_name = 0
    hostname = 0
    hw_type = 0
    tem_line1 = 0
    tem_line2 = 0
    pid_line1 = 0
    pid_line2 = 0
    tem_num = 0
    high_lcnum = 0
    high_lcpid = 0
    compare_result = 0
    lc_threshold = 0
    used_slot = 0
    used_lc_result = 0
    lc_slot = list()
    lc_tem = list()
    lc_name = list()
    thnf = 95
    thes = 86

    log_file = open("%s" %file_list[file_num], "r", encoding="UTF8")
    log_line = log_file.readlines()
    log_file.close()

    ### OS ###
    # show version 기반으로 해당 라우터의 운영체제를 검색 함.
    for os_parsing in range(len(log_line)):
        os_pattern = r'^Cisco\s(IOS|Internetwork Operating System|IOS XR)\sSoftware'
        os_result = re.search(os_pattern, log_line[os_parsing])
        if os_result != None:
            os_name = os_result.group()
            break
        else:
            pass

    ### Hostname ###
    # ios-xr일 경우 해당 if 에 match 됨.
    if os_name == "Cisco IOS XR Software":
        for host_parsing in range(len(log_line)):
            host_pattern = r'^\w\w/\d/\w\w\d/\w\w\w\d:.*#show version brief$' #show version brief 라는 문자열을 매치해서 hostname을 검색
            host_result = re.search(host_pattern, log_line[host_parsing])
            if host_result != None:
                hostname = host_result.group()
                hostname = hostname.split(sep = ":", maxsplit = 1) # : 문자를 기준으로 하나의 문자열을 두개로 나눔
                hostname = hostname[1] # 나눠진 두개의 문자열 중 2번째 문자열 선택
                hostname = hostname.split(sep = "#", maxsplit = 1) # 선택한 문자열을 # 기준으로 두개로 나눔
                hostname = hostname[0] # 나눠진 두개의 문자열 중 첫번째 문자열 선택
                hw_type = "ASR-9912" # 교환망은 ios-xr을 사용하는 라우터가 ASR-9912 밖에 없어서 hw-type는 수동으로 입력
                break # show version brief 라는 문자열 검색에 성공하면 문자열 검색 중지
            else:
                pass # show version brief 라는 문자열을 찾을 때 까지 반복

        if os.path.exists("%s\%s.txt" %(str(nextdir),hostname)) == True: # 로그 파일 중복 검색(만약 동일한 장비의 로그가 2개 있으면 해당 조건 match)
            if file_list[file_num] != "%s.txt" %hostname: # 현재 코드에서 open 된 log 파일의 이름이 hostname이 아니라면 
                os.remove("%s\%s" %(str(nextdir),file_list[file_num])) # 기존에 hostname으로 저장 된 log 파일 삭제
                print("####### remove %s(%s) #######" %(hostname, file_list[file_num])) # 현재 코드에서 open 된 log 파일의 이름을 hostname으로 변경 
        else: # 로그 파일이 중복 된 상태가 아니라면
            os.rename("%s" %file_list[file_num],"%s.txt" %hostname) # 현재 코드에서 open 된 log 파일의 이름을 hostname으로 변경

        #os.rename("%s" %file_list[file_num],"%s.txt" %hostname)
        print("Filename %s -> %s.txt" %(file_list[file_num], hostname))

    else: #ios-xr 을 제외한 나머지 장비는 해당 조건에 match (C6509 / C7609 / etc..)
        for host_parsing in range(len(log_line)):
            host_pattern = r'^.*#show version(|\s*)$' 
            host_result = re.search(host_pattern, log_line[host_parsing])
            if host_result != None:
                hostname = host_result.group()
                hostname = hostname.split(sep = "#s", maxsplit = 1)
                hostname = hostname[0]
                break
            else:
                pass

        if os.path.exists("%s\%s.txt" %(str(nextdir),hostname)) == True:
            if file_list[file_num] != "%s.txt" %hostname:
                os.remove("%s\%s" %(str(nextdir),file_list[file_num]))
                print("####### remove %s(%s) #######" %(hostname, file_list[file_num]))
        else:
            os.rename("%s" %file_list[file_num],"%s.txt" %hostname)

        #os.rename("%s" %file_list[file_num],"%s.txt" %hostname)
        print("Filename %s -> %s.txt" %(file_list[file_num], hostname))

        for hw_parsing in range(len(log_line)):
            hw_pattern = r'^PID:\s.*\s*.*$' #장비의 hw-type을 검색하기 위한 조건 (ex. WS-C6509E / WS-C7609S / etc..)
            hw_result = re.search(hw_pattern, log_line[hw_parsing])
            if hw_result != None:
                hw_type = hw_result.group()
                hw_type = hw_type.split(sep = " ", maxsplit = 2) # 공백을 기준으로 문자열을 3개로 나눔.
                hw_type = hw_type[1] # 3개로 나뉜 문자열 중 2번째 문자열 선택
                break
            else:
                pass

### Part. ASR9912 ###    

    if hw_type == "ASR-9912": # hw-type이 asr9912 인 장비들의 온도 조사를 위한 조건문
        for tem_start1 in range(len(log_line)):
            tem_pattern1 = r'^\w\w/\d/\w\w\d/\w\w\w\d:.*#show environment temperatures$' # 로그 중 온도 조사를 위해 확인이 필요한 부분을 발췌하기 위해 첫번째 열 지정 
            tem_result1 = re.search(tem_pattern1, log_line[tem_start1])
            if tem_result1 != None:
                tem_line1 = tem_start1
                break
            else:
                pass

        for tem_start2 in range(tem_line1 +1, len(log_line)):
            tem_pattern2 = r'^\w\w/\d/\w\w\d/\w\w\w\d:.*#' # 로그 중 온도 조사를 위해 확인이 필요한 부분을 발췌하기 위해 마지막 열 지정 
            tem_result2 = re.search(tem_pattern2, log_line[tem_start2])
            if tem_result2 != None:
                tem_line2 = tem_start2
                break
            else:
                pass

        for search_lc in range(tem_line1, tem_line2): # 위에서 지정한 첫 번째 > 마지막 열 까지 로그 검색을 실행
            lc_pattern = r'^\d/\d/.$' # 라인카드 번호 확인을 위해 검색 (ex. 0/5/CPU0)
            lc_result = re.search(lc_pattern, log_line[search_lc])
            if lc_result != None:
                lc_list = lc_result.group()
                lc_list = lc_list.split(sep = "/", maxsplit = 2) # 0/5/CPU0에서 "5"만 걸러내기 위한 split 구문
                lc_slot.append(lc_list[1]) # lc_slot 리스트에 현재 실장 된 라인카드의 번호를 순차적으로 저장
            else:
                pass

        for search_tem in range(tem_line1, tem_line2):
            lctem_pattern = r'^\s*host\s*Hotspot0\s*\d\d.\d$' # 라인카드 별 실시간 온도 조사 및 검색
            lctem_result = re.search(lctem_pattern, log_line[search_tem])
            if lctem_result != None: 
                lctem_list = lctem_result.group()
                lctem_pattern = r'\d{2,3}.\d' #온도 검색 (소수점 1자리 까지)
                lctem_result2 = re.search(lctem_pattern, lctem_list)
                lctem_list = lctem_result2.group() # lctem_list에 split 된 온도를 저장
                #lctem_list = lctem_list.split(sep = " ")
                #lctem_list = lctem_list[28]
                if tem_num <= 1:
                    pass
                else:
                    lc_tem.append(lctem_list)
                tem_num = tem_num + 1
            else:
                pass
        
        for pid_start in range(len(log_line)):
            pid_pattern = r'^\w*/\d/\w*\d/\w*\d:.*#admin show (plat|platform)$'
            pid_result = re.search(pid_pattern, log_line[pid_start])
            if pid_result != None:
                pid_line1 = pid_start
                break
            else:
                pass
        
        for pid_stop in range(pid_line1 + 1, len(log_line)):
            pid_pattern = r'^\w*/\d/\w*\d/\w*\d:.*#'
            pid_result = re.search(pid_pattern, log_line[pid_stop])
            if pid_result != None:
                pid_line2 = pid_stop
                break
            else:
                pass

        for search_pid in range(pid_line1, pid_line2):
            pid_pattern = r'^\d/\d/\w*\d\s*\S*'
            pid_result = re.search(pid_pattern, log_line[search_pid])
            if pid_result != None:
                pid_result2 = pid_result.group()
                pid_pattern = r'\w+-\w+-\w*'
                pid_name = re.search(pid_pattern, pid_result2) 
                pid_name = pid_name.group() 
                lc_name.append(pid_name)          
            else:
                pass            

        for tem_compare in range(0,len(lc_tem)):
            if str(compare_result) < lc_tem[tem_compare]:
                compare_result = lc_tem[tem_compare]
                high_lcnum = lc_slot[tem_compare]
                high_lcpid = lc_name[tem_compare]

        #print(hostname)
        #print("%s %s %s" %(high_lcnum, high_lcpid, compare_result))

### Part. C4510 ###

    elif hw_type == "WS-C4510R" or hw_type == "WS-C4006" or hw_type == "WS-C4506":
        for search_tem in range(len(log_line)):
            tem_pattern = r'^Chassis Temperature\s*=\s\d*\sdegrees\sCelsius'
            tem_result = re.search(tem_pattern, log_line[search_tem])
            if tem_result != None:
                tem_pattern = r'\d\d'
                tem_result2 = tem_result.group()
                tem_regex = re.search(tem_pattern, tem_result2)
                compare_result = tem_regex.group()
                high_lcnum = "Chassis"
                high_lcpid = "Chassis"
                break
            else:
                pass
        
        if hw_type == "WS-C4510R":
            for search_slot in range(len(log_line)):
                module_pattern = r'^\s+\d+\s+\d+\s+((\w+|\w+\W\w+)\s)+(\W\w+\W)*\s+'
                module_result = re.search(module_pattern, log_line[search_slot])
                if module_result != None:
                    module_pattern = r'^\s[^12]'
                    module_result = re.search(module_pattern, module_result.group())
                    if module_result != None:
                        module_result = module_result.group()
                        module_result = module_result.strip()
                        lc_slot.append(module_result)
                    else:
                        pass
                else:
                    pass

        else:
            for search_slot in range(len(log_line)):
                module_pattern = r'^\s{1,2}\d\s+\d{1,2}\s+((\w+|\w+\W\w+)\s)+(\W\w+\W)*\s+'
                module_result = re.search(module_pattern, log_line[search_slot])
                if module_result != None:
                    module_pattern = r'^\s[^1]'
                    module_result = re.search(module_pattern, module_result.group())
                    if module_result != None:
                        module_result = module_result.group()
                        module_result = module_result.strip()
                        lc_slot.append(module_result)
                    else:
                        pass
                else:
                    pass

        #print(hostname)
        #print("%s %s %s" %(high_lcnum, high_lcpid, compare_result))

### Part. C3945 ###
    
    elif hw_type == "CISCO3945-CHASSIS":
        for search_tem in range(len(log_line)):
            tem_pattern = r'^\s+CPU\stemperature:\s\d{2,3}\sCelsius'
            tem_result = re.search(tem_pattern, log_line[search_tem])
            if tem_result != None:
                tem_pattern = r'\d{2,3}'
                tem_result2 = tem_result.group()
                tem_regex = re.search(tem_pattern, tem_result2)
                compare_result = tem_regex.group()
                high_lcnum = "Chassis"
                high_lcpid = "Chassis"
                break
            else:
                pass

### Part. C6509 & C7609 ###

    else:
        for search_pid in range(len(log_line)):
            pid_pattern = r'^\s+[^56]\s+\d+\s+.*\w+-\w+-\w+\s+\w+$'
            pid_result = re.search(pid_pattern, log_line[search_pid])
            if pid_result != None:
                pid_result2 = pid_result.group()
                pid_pattern = r'\w+-\w+-\w+'
                pid_name = re.search(pid_pattern, pid_result2) 
                pid_name = pid_name.group()
                lc_name.append(pid_name)
            else:
                pass            

        for search_tem in range(len(log_line)):
            tem_pattern = r'^module [^56] outlet temperature: \d\dC$'
            tem_result = re.search(tem_pattern, log_line[search_tem])
            if tem_result != None:
                tem_result2 = tem_result.group()
                num_pattern = r'\d'
                tem_pattern = r'\d{2,3}'
                num_regex = re.search(num_pattern, tem_result2)
                tem_regex = re.search(tem_pattern, tem_result2)
                num_regex = num_regex.group()
                tem_regex = tem_regex.group()
                lc_slot.append(num_regex)
                lc_tem.append(tem_regex)
            else:
                pass

        if len(lc_slot) > 1:
            for tem_compare in range(0,len(lc_tem)):
                if str(compare_result) < lc_tem[tem_compare]:
                    compare_result = lc_tem[tem_compare]
                    high_lcnum = lc_slot[tem_compare]
                    high_lcpid = lc_name[tem_compare]
        else:
            high_lcnum = lc_slot[0]
            high_lcpid = lc_name[0]     
            compare_result = lc_tem[0]   

        for search_thres in range(len(log_line)):
            thres_pattern = r'^\s+threshold #2 for module %s outlet temperature:\s$' %high_lcnum
            thres_result = re.search(thres_pattern, log_line[search_thres])
            if thres_result != None:
                thres_pattern = r'\d+'
                thres_result2 = re.search(thres_pattern, log_line[search_thres + 1])
                lc_threshold = thres_result2.group()
                break
            else:
                pass

        #print(hostname)
        #print("%s %s %s %s" %(high_lcnum, high_lcpid, compare_result, lc_threshold))
    if len(lc_slot) > 1:
        for append_lc in range(len(lc_slot)):
            used_slot = "%s %s," %(used_slot, lc_slot[append_lc])
        used_lc_pattern = r'(\d, )+\d'
        used_lc_result = re.search(used_lc_pattern, used_slot)
        if used_lc_result != None:
            used_lc_result = used_lc_result.group()
            
        else:
            pass
    
    elif hw_type == "CISCO3945-CHASSIS":
        used_lc_result = "No Slots"

    else:
        used_lc_result = lc_slot[0]

    #print(hostname)
    #print(used_lc_result)

    wb = load_workbook("%s\SKB VoIP 온도 조사.xlsx" %nowdir)
    ws = wb["카드별 온도점검"]
    for excel_row in range(2,29):
        if ws["D%s" %excel_row].value == hostname:
            if hw_type == "WS-C6509-E" or hw_type == "CISCO7609-S" or hw_type == "WS-C6506-E":
                ws["E%s" %excel_row] = high_lcnum
                ws["F%s" %excel_row] = compare_result
                ws["H%s" %excel_row] = lc_threshold
                ws["I%s" %excel_row] = used_lc_result
                ws["J%s" %excel_row] = high_lcpid
            else:
                ws["E%s" %excel_row] = high_lcnum
                ws["F%s" %excel_row] = compare_result
                ws["I%s" %excel_row] = used_lc_result
                ws["J%s" %excel_row] = high_lcpid
                if hw_type == "WS-C4510R" or hw_type == "WS-C4006" or hw_type == "WS-C4506":
                    ws["H%s" %excel_row] = "95"
                elif hw_type == "CISCO3945-CHASSIS":
                    ws["H%s" %excel_row] = "90"
                else:
                    if high_lcpid == "A9K-24x10GE-TR" or high_lcpid == "A9K-40GE-TR" or high_lcpid == "A99-12x100GE" or high_lcpid == "A9K-MOD400-TR":
                        ws["H%s" %excel_row] = thnf
                    elif high_lcpid == "A9K-24X10GE-TR" or high_lcpid == "A99-12X100GE" or high_lcpid == "A9K-24x10GE-SE" or high_lcpid == "A9K-24X10GE-SE" :
                        ws["H%s" %excel_row] = thnf
                    elif high_lcpid == "A9K-24x10GE-1G-TR" or high_lcpid == "A9K-48x10GE-1G-TR":
                        ws["H%s" %excel_row] = thes
                    elif high_lcpid == "A9K-24X10GE-1G-TR" or high_lcpid == "A9K-48X10GE-1G-TR":
                        ws["H%s" %excel_row] = thes

        else:
            pass
    
    wb.save("%s\SKB VoIP 온도 조사.xlsx" %nowdir)

print("### Inspection Successful ###")
input("Press enter to exit ;)")


#Warning Threshold#
#show enviroment table

#STT(Shutdown) Threshold#
#show diag 0/x/CPU0 eeprom-info
